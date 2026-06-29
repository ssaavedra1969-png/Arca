#!/usr/bin/env python3
"""
Script de migracion de datos desde Excel/CSV a la base de datos de ARCA Facturador.
Soporta importacion de clientes, productos y facturas.
"""
import sys
import csv
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

from database.models import get_session_factory, init_db
from database.repositories import (
    ContribuyenteRepository, ProductoRepository, FacturaRepository
)
from config import ConfigManager
from loguru import logger


class Migrador:
    def __init__(self, db_path: str):
        self.session_factory = init_db(db_path)

    def migrar_clientes_csv(self, filepath: str) -> int:
        session = self.session_factory()
        try:
            repo = ContribuyenteRepository(session)
            count = 0
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    cuit = row.get("cuit", "").replace("-", "").replace(" ", "")
                    if not cuit or len(cuit) != 11:
                        logger.warning(f"CUIT invalido: {cuit}, saltando...")
                        continue
                    try:
                        repo.create({
                            "cuit": cuit,
                            "razon_social": row.get("razon_social", row.get("nombre", "")),
                            "condicion_iva": int(row.get("condicion_iva", 5)),
                            "domicilio": row.get("domicilio", ""),
                            "localidad": row.get("localidad", ""),
                            "provincia": row.get("provincia", ""),
                            "telefono": row.get("telefono", ""),
                            "email": row.get("email", ""),
                        })
                        count += 1
                    except Exception as e:
                        logger.error(f"Error importando {cuit}: {e}")
            session.commit()
            logger.info(f"Migrados {count} clientes desde {filepath}")
            return count
        finally:
            session.close()

    def migrar_clientes_excel(self, filepath: str) -> int:
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl no instalado. pip install openpyxl")
            return 0

        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        session = self.session_factory()
        try:
            repo = ContribuyenteRepository(session)
            count = 0
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                cuit = str(data.get("cuit", "")).replace("-", "").replace(" ", "")
                if not cuit or len(cuit) != 11:
                    continue
                try:
                    repo.create({
                        "cuit": cuit,
                        "razon_social": data.get("razon_social", data.get("nombre", "")),
                        "condicion_iva": int(data.get("condicion_iva", 5)),
                        "domicilio": data.get("domicilio", ""),
                    })
                    count += 1
                except Exception as e:
                    logger.error(f"Error: {e}")
            session.commit()
            return count
        finally:
            session.close()

    def migrar_productos_csv(self, filepath: str) -> int:
        session = self.session_factory()
        try:
            repo = ProductoRepository(session)
            count = 0
            with open(filepath, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        repo.create({
                            "codigo": row.get("codigo", f"IMP-{count}"),
                            "descripcion": row.get("descripcion", ""),
                            "tipo": row.get("tipo", "PRODUCTO"),
                            "precio_base": float(row.get("precio_base", 0)),
                            "alicuota_iva": float(row.get("alicuota_iva", 21)),
                        })
                        count += 1
                    except Exception as e:
                        logger.error(f"Error: {e}")
            session.commit()
            return count
        finally:
            session.close()

    def limpiar_datos_duplicados(self):
        session = self.session_factory()
        try:
            from sqlalchemy import func
            dups = session.query(
                ContribuyenteRepository.model_class.cuit,
                func.count(ContribuyenteRepository.model_class.id)
            ).group_by(ContribuyenteRepository.model_class.cuit).having(
                func.count(ContribuyenteRepository.model_class.id) > 1
            ).all()

            for cuit, count in dups:
                logger.warning(f"CUIT duplicado: {cuit} ({count} ocurrencias)")
            return len(dups)
        finally:
            session.close()


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Migrar datos a ARCA Facturador")
    parser.add_argument("tipo", choices=["clientes", "productos", "limpiar"],
                        help="Tipo de datos a migrar")
    parser.add_argument("archivo", nargs="?", help="Archivo CSV/Excel")
    parser.add_argument("--db", help="Ruta a la base de datos", default="data/arca_facturador.db")

    args = parser.parse_args()
    migrador = Migrador(args.db)

    if args.tipo == "limpiar":
        count = migrador.limpiar_datos_duplicados()
        print(f"Duplicados encontrados: {count}")
    elif args.archivo:
        if args.archivo.endswith(".csv"):
            if args.tipo == "clientes":
                count = migrador.migrar_clientes_csv(args.archivo)
            else:
                count = migrador.migrar_productos_csv(args.archivo)
        elif args.archivo.endswith((".xlsx", ".xls")):
            count = migrador.migrar_clientes_excel(args.archivo)
        else:
            print("Formato no soportado. Use .csv o .xlsx")
            return
        print(f"Migrados {count} registros de {args.tipo}")
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
