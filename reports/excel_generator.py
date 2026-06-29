from typing import Optional, List, Dict, Any
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from loguru import logger

from utils.formatters import Formatters


class ExcelGenerator:
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ROW_FONT = Font(name="Calibri", size=10)
    TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    THIN_BORDER = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    def generar_historial(self, facturas: List[Dict[str, Any]],
                          output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial Facturas"

        ws.merge_cells("A1:K1")
        ws["A1"] = "Historial de Facturas Electronicas"
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:K2")
        ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = [
            "Fecha", "Tipo", "Pto.Venta", "Numero", "Cliente",
            "CUIT", "Subtotal", "IVA", "Total", "CAE", "Estado"
        ]
        self._write_header(ws, 4, headers)

        for i, fac in enumerate(facturas):
            row = i + 5
            ws.cell(row=row, column=1, value=Formatters.formatear_fecha(fac.get("fecha_emision")))
            ws.cell(row=row, column=2, value=Formatters.tipo_comprobante_descripcion(fac.get("tipo_comprobante", 0)))
            ws.cell(row=row, column=3, value=fac.get("punto_venta", 0))
            ws.cell(row=row, column=4, value=fac.get("numero_factura", 0))
            ws.cell(row=row, column=5, value=fac.get("razon_social_cliente", ""))
            ws.cell(row=row, column=6, value=Formatters.formatear_cuit(fac.get("cuit_cliente", "")))
            ws.cell(row=row, column=7, value=fac.get("subtotal", 0))
            ws.cell(row=row, column=8, value=fac.get("iva_total", 0))
            ws.cell(row=row, column=9, value=fac.get("total", 0))
            ws.cell(row=row, column=10, value=Formatters.formatear_cae(fac.get("cae", "")))
            ws.cell(row=row, column=11, value=Formatters.estado_comprobante(fac.get("estado", "")))

            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.font = self.ROW_FONT
                cell.border = self.THIN_BORDER
                if col in (7, 8, 9):
                    cell.number_format = '$ #,##0.00'
                    cell.alignment = Alignment(horizontal="right")

            if i % 2 == 1:
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = self.ALT_FILL

        col_widths = [14, 20, 10, 10, 35, 18, 14, 14, 14, 18, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.auto_filter.ref = f"A4:K{len(facturas) + 4}"

        wb.save(output_path)
        logger.info(f"Excel generado: {output_path}")
        return output_path

    def generar_reporte_mensual(self, data: List[Dict[str, Any]],
                                output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Mensual"

        ws.merge_cells("A1:E1")
        ws["A1"] = "Reporte Mensual de Facturacion"
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = ["Mes", "Cant Facturas", "Subtotal", "IVA", "Total"]
        self._write_header(ws, 3, headers)

        for i, row_data in enumerate(data):
            row = i + 4
            for j, val in enumerate(row_data):
                ws.cell(row=row, column=j + 1, value=val)
                cell = ws.cell(row=row, column=j + 1)
                cell.font = self.ROW_FONT
                cell.border = self.THIN_BORDER
                if j >= 2:
                    cell.number_format = '$ #,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                if i % 2 == 1:
                    cell.fill = self.ALT_FILL

        col_widths = [15, 15, 18, 18, 18]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(output_path)
        return output_path

    def _write_header(self, ws, row: int, headers: List[str]):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER
        ws.row_dimensions[row].height = 25


def exportar_historial_excel(facturas: List[Dict[str, Any]], output_path: str) -> str:
    gen = ExcelGenerator()
    return gen.generar_historial(facturas, output_path)
